from testTypesCode import assertionRoulette, conditionalTestLogic, duplicateAssert, emptyTest, magicNumber, redundantPrint, redundantAssertion, unknownTest
from testTypesCode import sensitiveEquality, sleepyTest, exceptionHandling, eagerTest, lazyTest, mysteryGuest, generalFixture, constructorInitialization
from testTypesCode import resourceOptimism
import glob, os
import openpyxl
import pandas as pd
import time

headers = [
    "",
    "AssertionRoulette", 
    "conditionalTestLogic", 
    "duplicateAssert", 
    "emptyTest", 
    "magicNumber", 
    "redundantPrint",
    "redundantAssertion",
    "unknownTest",
    "sensitiveEquality",
    "sleepyTest",
    "exceptionHandling",
    "eagerTest",
    "lazyTest",
    "mysteryGuest",
    "generalFixture",
    "constructorInitialization",
    "resourceOptimism"
    ]

def runAssertionForAllFiles(allFiles, repo, allFilesSheet):
    assertionRouletteResponses = []
    conditionalTestLogicResponses = []
    duplicateAssertResponses = []
    emptyTestResponses = []
    magicNumberResponses = []
    redundantPrintResponses = []
    redundantAssertionResponses = []
    unknownTestResponses = []
    sensitiveEqualityResponses = []
    sleepyTestResponses = []
    exceptionHandlingResponses = []
    eagerTestResponses = []
    lazyTestResponses = []
    mysteryGuestResponses = []
    generalFixtureResponses = []
    constructorInitializationResponses = []
    resourceOptimismResponses = []
    response = []

    workbookFile = openpyxl.Workbook()
    sheetFile = workbookFile.active
    sheetFile.append(headers)
    if len(allFiles) == 0:
        return
    for singleFile in allFiles:
        print(singleFile)
        path = os.path.basename(os.path.normpath(singleFile))
        xmlPath = repo + "XMLCode/" + path + ".xml"
        os.system("srcml " + singleFile + " -o " + xmlPath)
        line = []
        lineFiles = [""]
        line.append(singleFile)
        assertionRouletterResponse = assertionRoulette.assertRule("./" + xmlPath)
        assertionRouletteResponses += assertionRouletterResponse
        if True not in assertionRouletterResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(assertionRouletterResponse))
        conditionalTestLogicResponse = conditionalTestLogic.assertRule("./" + xmlPath)
        conditionalTestLogicResponses += conditionalTestLogicResponse
        if True not in conditionalTestLogicResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(conditionalTestLogicResponse))
        duplicateAssertResponse = duplicateAssert.assertRule("./" + xmlPath)
        duplicateAssertResponses += duplicateAssertResponse
        if True not in duplicateAssertResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(duplicateAssertResponse))
        emptyTestResponse = emptyTest.assertRule("./" + xmlPath)
        emptyTestResponses += emptyTestResponse
        if True not in emptyTestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(emptyTestResponse))
        magicNumberResponse = magicNumber.assertRule("./" + xmlPath)
        magicNumberResponses += magicNumberResponse
        if True not in magicNumberResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(magicNumberResponse))
        redundantPrintResponse = redundantPrint.assertRule("./" + xmlPath)
        redundantPrintResponses += redundantPrintResponse
        if True not in redundantPrintResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(redundantPrintResponse))
        redundantAssertionResponse = redundantAssertion.assertRule("./" + xmlPath)
        redundantAssertionResponses += redundantAssertionResponse
        if True not in redundantAssertionResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(redundantAssertionResponse))
        unknownTestResponse = unknownTest.assertRule("./" + xmlPath)
        unknownTestResponses += unknownTestResponse
        if True not in unknownTestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(unknownTestResponse))
        sensitiveEqualityResponse = sensitiveEquality.assertRule("./" + xmlPath)
        sensitiveEqualityResponses += sensitiveEqualityResponse
        if True not in sensitiveEqualityResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(sensitiveEqualityResponse))
        sleepyTestResponse = sleepyTest.assertRule("./" + xmlPath)
        sleepyTestResponses += sleepyTestResponse
        if True not in sleepyTestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(sleepyTestResponse))
        exceptionHandlingResponse = exceptionHandling.assertRule("./" + xmlPath)
        exceptionHandlingResponses += exceptionHandlingResponse
        if True not in exceptionHandlingResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(exceptionHandlingResponse))
        eagerTestResponse = eagerTest.assertRule("./" + xmlPath)
        eagerTestResponses += eagerTestResponse
        if True not in eagerTestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(eagerTestResponse))
        lazyTestResponse = lazyTest.assertRule("./" + xmlPath)
        lazyTestResponses += lazyTestResponse
        if True not in lazyTestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(lazyTestResponse))
        mysteryGuestResponse = mysteryGuest.assertRule("./" + xmlPath)
        mysteryGuestResponses += mysteryGuestResponse
        if True not in mysteryGuestResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(mysteryGuestResponse))
        generalFixtureResponse = generalFixture.assertRule("./" + xmlPath)
        generalFixtureResponses += generalFixtureResponse
        if True not in generalFixtureResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(generalFixtureResponse))
        constructorInitializationResponse = constructorInitialization.assertRule("./" + xmlPath)
        constructorInitializationResponses += constructorInitializationResponse
        if True not in constructorInitializationResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(constructorInitializationResponse))
        resourceOptimismResponse = resourceOptimism.assertRule("./" + xmlPath)
        resourceOptimismResponses += resourceOptimismResponse
        if True not in resourceOptimismResponse:
            lineFiles.append('NA')
        else:
            lineFiles.append(singleFile)
        line.append(len(resourceOptimismResponse))
        sheetFile.append(line)
        allFilesSheet.append(lineFiles)
    workbookFile.save(repo.replace("repo/", "") + ".xlsx")
    print("AssertionRoulette: " + str(assertionRouletteResponses.count(True)/len(assertionRouletteResponses)))
    print("conditionalTestLogicResponses: " + str(conditionalTestLogicResponses.count(True)/len(conditionalTestLogicResponses)))
    print("duplicateAssertResponses: " + str(duplicateAssertResponses.count(True)/len(duplicateAssertResponses)))
    print("emptyTestResponses: " + str(emptyTestResponses.count(True)/len(emptyTestResponses)))
    print("magicNumberResponses: " + str(magicNumberResponses.count(True)/len(magicNumberResponses)))
    print("redundantPrintResponses: " + str(redundantPrintResponses.count(True)/len(redundantPrintResponses)))
    print("redundantAssertionResponses: " + str(redundantAssertionResponses.count(True)/len(redundantAssertionResponses)))
    print("unknownTestResponses: " + str(unknownTestResponses.count(True)/len(unknownTestResponses)))
    print("sensitiveEqualityResponses: " + str(sensitiveEqualityResponses.count(True)/len(sensitiveEqualityResponses)))
    print("sleepyTestResponses: " + str(sleepyTestResponses.count(True)/len(sleepyTestResponses)))
    print("exceptionHandlingResponses: " + str(exceptionHandlingResponses.count(True)/len(exceptionHandlingResponses)))
    print("eagerTestResponses: " + str(eagerTestResponses.count(True)/len(eagerTestResponses)))
    print("lazyTestResponses: " + str(lazyTestResponses.count(True)/len(lazyTestResponses)))
    print("mysteryGuestResponses: " + str(mysteryGuestResponses.count(True)/len(mysteryGuestResponses)))
    print("generalFixtureResponses: " + str(generalFixtureResponses.count(True)/len(generalFixtureResponses)))
    print("constructorInitializationResponses: " + str(constructorInitializationResponses.count(True)/len(constructorInitializationResponses)))
    print("resourceOptimismResponses: " + str(resourceOptimismResponses.count(True)/len(resourceOptimismResponses)))
    print("-----------------------------------------")
    # response.append(assertionRouletteResponses.count(True)/len(assertionRouletteResponses))
    # response.append(conditionalTestLogicResponses.count(True)/len(conditionalTestLogicResponses))
    # response.append(duplicateAssertResponses.count(True)/len(duplicateAssertResponses))
    # response.append(emptyTestResponses.count(True)/len(emptyTestResponses))
    # response.append(magicNumberResponses.count(True)/len(magicNumberResponses))
    # response.append(redundantPrintResponses.count(True)/len(redundantPrintResponses))
    # response.append(redundantAssertionResponses.count(True)/len(redundantAssertionResponses))
    # response.append(unknownTestResponses.count(True)/len(unknownTestResponses))
    # response.append(sensitiveEqualityResponses.count(True)/len(sensitiveEqualityResponses))
    # response.append(sleepyTestResponses.count(True)/len(sleepyTestResponses))
    # response.append(exceptionHandlingResponses.count(True)/len(exceptionHandlingResponses))
    # response.append(eagerTestResponses.count(True)/len(eagerTestResponses))
    # response.append(lazyTestResponses.count(True)/len(lazyTestResponses))
    # response.append(mysteryGuestResponses.count(True)/len(mysteryGuestResponses))
    # response.append(generalFixtureResponses.count(True)/len(generalFixtureResponses))
    # response.append(constructorInitializationResponses.count(True)/len(constructorInitializationResponses))
    # response.append(resourceOptimismResponses.count(True)/len(resourceOptimismResponses))
    response.append(assertionRouletteResponses.count(True))
    response.append(conditionalTestLogicResponses.count(True))
    response.append(duplicateAssertResponses.count(True))
    response.append(emptyTestResponses.count(True))
    response.append(magicNumberResponses.count(True))
    response.append(redundantPrintResponses.count(True))
    response.append(redundantAssertionResponses.count(True))
    response.append(unknownTestResponses.count(True))
    response.append(sensitiveEqualityResponses.count(True))
    response.append(sleepyTestResponses.count(True))
    response.append(exceptionHandlingResponses.count(True))
    response.append(eagerTestResponses.count(True))
    response.append(lazyTestResponses.count(True))
    response.append(mysteryGuestResponses.count(True))
    response.append(generalFixtureResponses.count(True))
    response.append(constructorInitializationResponses.count(True))
    response.append(resourceOptimismResponses.count(True))
    return response

def getOnly4Digits(value):
    return '%.4f'%value

def getPercentage(value):
    return ('%.2f'%(value*100) + '%')

start_time = time.time()
workbook = openpyxl.Workbook()
workbook2 = openpyxl.Workbook()
allFilesWorkbook = openpyxl.Workbook()
sheet = workbook.active
sheet2 = workbook2.active
allFilesSheet = allFilesWorkbook.active
sheet.append(headers)
sheet2.append(headers)
allFilesSheet.append(headers)

allRepos = glob.glob("repos/*")
for repo in allRepos:
    print(repo)
    if "XMLCode" in repo:
        continue
    os.system("mkdir ./" + repo + "XMLCode")
    allFiles = glob.glob(repo + "/**/*test*/**/*.c", recursive=True)
    response = runAssertionForAllFiles(allFiles, repo, allFilesSheet)
    if response is not None:
        response4Digits = map(getOnly4Digits, response)
        responsePercentage = map(getPercentage, response)
        sheet.append([repo] + list(response4Digits))
        sheet2.append([repo] + list(responsePercentage))
workbook.save("results.xlsx")
workbook2.save("results%.xlsx")
allFilesWorkbook.save("filesWithSmells.xlsx")
print("--- %s seconds ---" % (time.time() - start_time))

